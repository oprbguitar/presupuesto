# -*- coding: utf-8 -*-
"""Ingesta REAL de adjudicaciones SEACE (OECE/CONOSCE) + presupuesto MEF.

Fuente oficial: reportes anuales de adjudicaciones del OECE (portal CONOSCE), enlazados desde
el tablero de datos abiertos de SEACE. Formato XLSX a nivel de ítem, ~19 MB/año, con:
  entidad + RUC + departamento, objeto contractual (incluye OBRAS), tipo de procedimiento,
  monto referencial y monto adjudicado, proveedor + RUC, y fechas de convocatoria/buena pro.

Cubre 2023, 2024 y 2025 (el consolidado del año en curso aún no se publica; para 2026 se
mantiene la ingesta de Perú Compras). El presupuesto (PIA/PIM/devengado/girado) se toma de
data/mef.js (ya presente en el portal) a nivel nacional, para el embudo presupuestal.

Reglas: el monto adjudicado NO equivale al devengado ni al girado (se guardan por separado).
"""
import argparse, io, json, os, re, sys, heapq, urllib.request, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from _comun import guardar_json, cargar_json, log, hoy, DATA_DIR, RAW_DIR

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120 Safari/537.36"}
TOP_FILAS = 5000
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

# ------- utilidades de red / resolución de URLs -------
def resolver_base(anio):
    """Resuelve la URL del XLSX de adjudicaciones del año vía el acortador oficial del tablero."""
    tag = "conosceadjudicaciones%d" % anio
    try:
        html = urllib.request.urlopen(urllib.request.Request(
            "https://tinyurl.com/preview/download/" + tag, headers=UA), timeout=60).read().decode("utf-8", "replace")
    except Exception as e:  # noqa: BLE001
        log("No se pudo resolver %s: %s" % (tag, e), "warn"); return None
    cands = [c for c in re.findall(r'https?://[^\s"\'<>]+', html) if "conosce.osce" in c and c.endswith(".xlsx")]
    return cands[0] if cands else None

def partes_de(base):
    urls = []
    for n in range(0, 12):
        u = re.sub(r"_\d+\.xlsx$", "_%d.xlsx" % n, base)
        try:
            r = urllib.request.urlopen(urllib.request.Request(u, headers=UA, method="HEAD"), timeout=40)
            if r.status == 200:
                urls.append(u); continue
        except Exception:
            break
        break
    return urls

def descargar(u, dst):
    if os.path.exists(dst) and os.path.getsize(dst) > 1000:
        return dst
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    urllib.request.urlretrieve(u, dst)  # nosec - fuente oficial
    return dst

# ------- normalización de campos -------
def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return round(float(v), 2)
    s = str(v).strip().replace(" ", "")
    if not s: return None
    if s.count(",") and s.count("."):
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    elif s.count(","):
        s = s.replace(",", ".")
    try: return round(float(s), 2)
    except ValueError: return None

def objeto_de(v):
    t = (v or "").strip().lower()
    if "obra" in t and "consult" in t: return "Consultoría de obras"
    if "consult" in t and "obra" in t: return "Consultoría de obras"
    if t.startswith("consultor") and "obra" in t: return "Consultoría de obras"
    if "obra" in t: return "Obras"
    if "consultor" in t: return "Consultoría de obras"
    if "bien" in t: return "Bienes"
    if "servicio" in t: return "Servicios"
    return (v or "Otros").strip().title()

def fecha_iso(f):
    if f is None: return None
    if isinstance(f, datetime.datetime): return f.strftime("%Y-%m-%d")
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(f))
    if m: return "%s-%02d-%02d" % (m.group(3), int(m.group(2)), int(m.group(1)))
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(f))
    return m.group(0) if m else None

def nivel_de(tipoentidad, nombre):
    t = (tipoentidad or "").upper() + " " + (nombre or "").upper()
    if "REGIONAL" in t: return "Regional"
    if "MUNICIPALIDAD" in t or "LOCAL" in t: return "Local"
    return "Nacional"

# ------- MEF (presupuesto nacional) desde data/mef.js -------
def mef_totales():
    try:
        txt = open(os.path.join(ROOT, "data", "mef.js"), encoding="utf-8").read()
        m = re.search(r"window\.MEF\s*=\s*(\{.*\})\s*;?\s*$", txt, re.S)
        obj = json.loads(m.group(1))
        out = {}
        for y, d in obj.get("years", {}).items():
            t = d.get("total") or {}
            out[int(y)] = {"pia": t.get("pia"), "pim": t.get("pim"), "dev": t.get("dev"), "gir": t.get("gir")}
        return out
    except Exception as e:  # noqa: BLE001
        log("No se pudo leer data/mef.js: %s" % e, "warn"); return {}

# ------- agregación -------
class Agg:
    def __init__(self):
        self.n_items = 0; self.procesos = set(); self.total_adj = 0.0; self.total_ref = 0.0
        self.provs = {}; self.ents = {}; self.deps = {}; self.tipos = {}; self.cats = {}
        self.obj = {}; self.desiertos = 0; self.anulados = 0; self.cd_n = 0; self.cd_m = 0.0
        self.n_oc = 0; self.n_os = 0; self.heap = []; self._c = 0

OBJ_ORDER = ["Bienes", "Servicios", "Obras", "Consultoría de obras"]

def procesar_xlsx(path, agg, anio):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip().lower() if h is not None else "" for h in next(it)]
    idx = {name: i for i, name in enumerate(hdr)}
    def g(row, name):
        i = idx.get(name); return row[i] if (i is not None and i < len(row)) else None
    for row in it:
        if row is None: continue
        ent = (g(row, "entidad") or "").strip()
        ruc_ent = str(g(row, "entidad_ruc") or "").strip()
        dep = (g(row, "entidad_departamento") or g(row, "departamento_item") or "").strip()
        obj = objeto_de(g(row, "objetocontractual"))
        tipo = (g(row, "tipoprocesoseleccion") or "").strip() or "Otro"
        estado = (g(row, "estado_item") or "").strip()
        proc = str(g(row, "codigoconvocatoria") or "").strip()
        prov = (g(row, "proveedor") or "").strip()
        ruc_prov = str(g(row, "ruc_proveedor") or "").strip()
        ref = num(g(row, "monto_referencial_item_soles"))
        adj = num(g(row, "monto_adjudicado_item_soles"))
        m = adj if adj is not None else 0.0

        agg.n_items += 1
        if proc: agg.procesos.add(proc)
        agg.total_adj += m
        agg.total_ref += ref or 0.0
        if obj == "Bienes": agg.n_oc += 1
        elif obj == "Servicios": agg.n_os += 1
        if re.search(r"desiert", estado, re.I): agg.desiertos += 1
        if re.search(r"nul|anulad", estado, re.I): agg.anulados += 1
        if re.search(r"directa", tipo, re.I): agg.cd_n += 1; agg.cd_m += m
        agg.obj[obj] = agg.obj.get(obj, [0.0, 0]); agg.obj[obj][0] += m; agg.obj[obj][1] += 1
        agg.tipos[tipo] = agg.tipos.get(tipo, [0.0, 0]); agg.tipos[tipo][0] += m; agg.tipos[tipo][1] += 1
        cat = obj  # categoría base = objeto (SEACE no trae categoría fina homogénea)
        agg.cats[cat] = agg.cats.get(cat, [0.0, 0]); agg.cats[cat][0] += m; agg.cats[cat][1] += 1
        if ruc_prov:
            p = agg.provs.setdefault(ruc_prov, {"nombre": prov, "monto": 0.0, "n": 0}); p["monto"] += m; p["n"] += 1
        if ruc_ent:
            e = agg.ents.setdefault(ruc_ent, {"nombre": ent, "nivel": nivel_de(g(row, "tipoentidad"), ent),
                                              "dep": dep, "monto": 0.0, "n": 0,
                                              "Bienes": 0.0, "Servicios": 0.0, "Obras": 0.0, "Consultoría de obras": 0.0})
            e["monto"] += m; e["n"] += 1
            if obj in e: e[obj] += m
        depN = dep.upper()
        if depN and depN not in ("NULL", "-", "SIN DEPARTAMENTO"):
            agg.deps[depN] = agg.deps.get(depN, [0.0, 0]); agg.deps[depN][0] += m; agg.deps[depN][1] += 1

        agg._c += 1
        fila = {"codigo": proc or ("SEACE-%s-%d" % (anio, agg._c)), "ruc_ent": ruc_ent, "entidad": ent,
                "objeto": obj, "categoria": (g(row, "descripcion_proceso") or obj), "tipo": tipo,
                "estado": estado or "Adjudicado", "proveedor": prov, "ruc": ruc_prov,
                "ref": ref, "adj": adj, "dep": dep, "nivel": nivel_de(g(row, "tipoentidad"), ent),
                "f_conv": fecha_iso(g(row, "fecha_convocatoria")), "f_adj": fecha_iso(g(row, "fecha_buenapro"))}
        item = (m, agg._c, fila)
        if len(agg.heap) < TOP_FILAS: heapq.heappush(agg.heap, item)
        elif m > agg.heap[0][0]: heapq.heapreplace(agg.heap, item)
    wb.close()

def construir(anio, agg, mef, parcial, ent_ids):
    def topd(d, n):
        return sorted(({"k": k, "monto": round(v[0], 2), "n": v[1]} for k, v in d.items()), key=lambda x: -x["monto"])[:n]
    top_prov = sorted(({"proveedor": v["nombre"], "ruc": k, "monto": round(v["monto"], 2), "n": v["n"]}
                       for k, v in agg.provs.items()), key=lambda x: -x["monto"])[:25]
    por_ent = sorted(({"id": ent_ids.get(k), "entidad": v["nombre"], "ruc": k, "nivel": v["nivel"],
                       "sector": "", "departamento": (v["dep"] or "").title(),
                       "convocado": round(v["monto"], 2), "adjudicado": round(v["monto"], 2),
                       "contratado": round(v["monto"], 2),
                       "bienes": round(v["Bienes"], 2), "servicios": round(v["Servicios"], 2),
                       "obras": round(v["Obras"], 2), "consultoria": round(v["Consultoría de obras"], 2),
                       "n": v["n"]}
                      for k, v in agg.ents.items()), key=lambda x: -x["adjudicado"])[:700]
    dist = [{"objeto": o, "monto": round(agg.obj.get(o, [0, 0])[0], 2), "n": agg.obj.get(o, [0, 0])[1]}
            for o in OBJ_ORDER if agg.obj.get(o)]
    adj = round(agg.total_adj, 2); ref = round(agg.total_ref, 2)
    b = mef.get(anio, {})
    agregados = {
        "anio": anio, "parcial": parcial, "actualizado": hoy(), "estado_datos": "validado",
        "fuente": {"nombre": "OECE/SEACE — Adjudicaciones (CONOSCE) + MEF (Consulta Amigable)",
                   "url": "https://bi.seace.gob.pe/",
                   "cobertura": "Ítems adjudicados en procedimientos de selección (incluye obras y consultorías). Presupuesto nacional del MEF.",
                   "metodologia": "Suma del monto adjudicado por ítem, agregado por entidad, objeto, tipo, departamento y proveedor. PIA/PIM/dev/girado del MEF (nivel nacional)."},
        "indicadores": {
            "pia": b.get("pia"), "pim": b.get("pim"), "dev": b.get("dev"), "gir": b.get("gir"),
            "pac_programado": None, "convocado": ref, "adjudicado": adj, "contratado": adj,
            "n_procedimientos": len(agg.procesos), "n_oc": agg.n_oc, "n_os": agg.n_os,
            "n_proveedores": len(agg.provs), "cd_num": agg.cd_n, "cd_monto": round(agg.cd_m, 2),
            "desiertos": agg.desiertos, "anulados": agg.anulados},
        "distribucion_objeto": dist,
        "top_categorias": [{"categoria": x["k"], "monto": x["monto"], "n": x["n"]} for x in topd(agg.cats, 15)],
        "top_proveedores": top_prov,
        "por_tipo_procedimiento": [{"tipo": x["k"], "monto": x["monto"], "n": x["n"]} for x in topd(agg.tipos, 30)],
        "comparacion_montos": {"convocado": ref, "adjudicado": adj, "contratado": adj},
        "embudo": [{"etapa": "PIM", "monto": b.get("pim")}, {"etapa": "PAC programado", "monto": None},
                   {"etapa": "Referencial", "monto": ref}, {"etapa": "Adjudicado", "monto": adj},
                   {"etapa": "Contratado", "monto": adj}, {"etapa": "Devengado", "monto": b.get("dev")}],
        "por_departamento": sorted(({"departamento": k.title(), "monto": round(v[0], 2), "n": v[1]}
                                    for k, v in agg.deps.items()), key=lambda x: -x["monto"]),
        "por_entidad": por_ent}
    guardar_json("agregados_%s.json" % anio, agregados)

    rows = []
    for _, _, f in sorted(agg.heap, key=lambda t: -t[0]):
        rows.append({"codigo": f["codigo"], "entidad_id": ent_ids.get(f["ruc_ent"]), "entidad": f["entidad"], "ue": "",
                     "descripcion": f["categoria"], "objeto": f["objeto"], "categoria": f["objeto"], "tipo": f["tipo"],
                     "estado": f["estado"], "proveedor": f["proveedor"], "ruc": f["ruc"], "sancionado": False,
                     "convocado": f["ref"], "adjudicado": f["adj"], "contratado": f["adj"],
                     "f_convocatoria": f["f_conv"], "f_adjudicacion": f["f_adj"], "postores": None,
                     "regimen": "Ley 30225 (Contrataciones del Estado)", "nivel": f["nivel"], "sector": "",
                     "departamento": (f["dep"] or "").title(), "ubigeo": "",
                     "cierre_anio": (f["f_conv"] or "")[5:7] == "12",
                     "url": "https://prodapp2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml"})
    guardar_json("procedimientos_%s.json" % anio, {
        "anio": anio, "parcial": parcial, "actualizado": hoy(), "estado_datos": "validado",
        "rows_tope": TOP_FILAS, "rows_total": agg.n_items, "fuente": agregados["fuente"], "rows": rows}, optimizado=True)
    return {"anio": anio, "pim": b.get("pim"), "convocado": ref, "adjudicado": adj,
            "contratado": adj, "dev": b.get("dev"), "parcial": parcial}

def main(anios, parcial):
    mef = mef_totales()
    ent_global = {}; prov_global = {}; aggs = {}
    for anio in anios:
        base = resolver_base(anio)
        if not base:
            log("SEACE %d: sin consolidado publicado; se omite (usar Perú Compras para ese año)." % anio, "warn")
            continue
        agg = Agg()
        for u in partes_de(base):
            dst = os.path.join(RAW_DIR, "seace_adj_%d_%s" % (anio, u.rsplit("_", 1)[-1]))
            log("SEACE %d: descargando %s…" % (anio, os.path.basename(u)))
            descargar(u, dst)
            log("SEACE %d: procesando %s…" % (anio, os.path.basename(dst)))
            procesar_xlsx(dst, agg, anio)
        log("SEACE %d: %d ítems, %d procesos, adjudicado=%.0f" % (anio, agg.n_items, len(agg.procesos), agg.total_adj))
        for ruc, v in agg.ents.items(): ent_global.setdefault(ruc, (v["nombre"], v["nivel"], v["dep"]))
        for ruc, v in agg.provs.items():
            g = prov_global.setdefault(ruc, [v["nombre"], 0.0]); g[1] += v["monto"]
        aggs[anio] = agg

    ent_ids = {ruc: "E-%04d" % (i + 1) for i, ruc in enumerate(sorted(ent_global))}
    evol = []
    for anio in sorted(aggs):
        evol.append(construir(anio, aggs[anio], mef, anio == parcial, ent_ids))

    # 2026 (u otros años sin SEACE): conservar lo existente (Perú Compras) e incluir en evolución
    # Años totales del módulo = SEACE + cualquier año con agregados ya presentes (p. ej. 2026 Perú Compras).
    anios_seace = set(aggs)
    todos = set(anios)
    for y in range(2023, parcial + 1):
        if os.path.exists(os.path.join(DATA_DIR, "agregados_%d.json" % y)):
            todos.add(y)
    todos = sorted(todos)
    for anio in todos:
        if anio in anios_seace: continue
        ag = cargar_json("agregados_%s.json" % anio)
        if ag:
            i = ag.get("indicadores", {})
            evol.append({"anio": anio, "pim": i.get("pim"), "convocado": i.get("convocado"),
                         "adjudicado": i.get("adjudicado"), "contratado": i.get("contratado"),
                         "dev": i.get("dev"), "parcial": anio == parcial})
    evol.sort(key=lambda x: x["anio"])

    # Maestro de entidades (real, con departamento de SEACE) + unión con lo existente (2026)
    maestro_prev = cargar_json("maestro_entidades.json", {"entidades": []}) or {"entidades": []}
    ents = []
    vistos = set()
    for ruc in sorted(ent_global):
        nombre, nivel, dep = ent_global[ruc]; vistos.add(ruc)
        ents.append({"id": ent_ids[ruc], "nombre": nombre, "ruc": ruc, "cod_oece": None, "cod_pliego_mef": None,
                     "cod_ue_mef": None, "sector": "", "nivel": nivel, "ubigeo": "", "departamento": (dep or "").title()})
    for e in maestro_prev.get("entidades", []):
        if e.get("ruc") and e["ruc"] not in vistos:
            ents.append(e); vistos.add(e["ruc"])
    guardar_json("maestro_entidades.json", {"actualizado": hoy(), "estado_datos": "validado",
        "nota": "Entidades de adjudicaciones SEACE (con departamento) y Perú Compras. No se asume correspondencia 1:1 con pliego/UE del MEF.",
        "entidades": ents})

    # Catálogo de proveedores acotado a los 2000 de mayor monto (el resto no lo usa la UI).
    top_prov_global = sorted(prov_global.items(), key=lambda kv: -kv[1][1])[:2000]
    guardar_json("proveedores.json", {"actualizado": hoy(), "estado_datos": "validado",
        "nota": "Top 2000 proveedores por monto adjudicado (catálogo acotado).",
        "proveedores": [{"ruc": ruc, "nombre": v[0], "monto": round(v[1], 2), "sancionado": False, "sancion_detalle": ""}
                        for ruc, v in top_prov_global]})

    manifest = cargar_json("manifest.json", {}) or {}
    manifest.update({
        "modulo": "contrataciones-publicas", "version": 3, "estado_datos": "validado", "actualizado": hoy(),
        "cobertura": "Adjudicaciones SEACE (procedimientos de selección, incluye obras) 2023-2025 + órdenes por "
                     "Catálogos Electrónicos de Perú Compras (2026). Presupuesto nacional del MEF.",
        "anios": todos, "anio_parcial": parcial, "evolucion": evol,
        "archivos": {"agregados": "agregados_{anio}.json", "procedimientos": "procedimientos_{anio}.json",
                     "maestro_entidades": "maestro_entidades.json", "proveedores": "proveedores.json",
                     "geo_departamentos": "peru_departamentos.json"},
        "geo_nota": "Departamento tomado del campo oficial de la entidad en SEACE.",
        "fuentes": [
            {"nombre": "OECE/SEACE — Adjudicaciones (CONOSCE)", "url": "https://bi.seace.gob.pe/",
             "descripcion": "Reportes anuales de adjudicaciones a nivel de ítem (incluye obras y consultorías)."},
            {"nombre": "MEF — Consulta Amigable", "url": "https://apps5.mineco.gob.pe/transparencia/Navegador/default.aspx",
             "descripcion": "PIA, PIM, devengado y girado (nivel nacional) para el embudo presupuestal."},
            {"nombre": "Perú Compras — Catálogos Electrónicos", "url": "https://www.datosabiertos.gob.pe/",
             "descripcion": "Órdenes por Acuerdos Marco (cobertura del año en curso)."},
            {"nombre": "Datos Abiertos del Estado Peruano", "url": "https://www.datosabiertos.gob.pe/",
             "descripcion": "Portal CKAN de conjuntos de datos oficiales."}],
        "limitaciones": [
            "El monto adjudicado (SEACE) no equivale al devengado ni al girado (MEF).",
            "El presupuesto MEF es nacional; no se atribuye a un procedimiento concreto.",
            "La tabla detallada muestra los %d ítems de mayor monto por año; los agregados usan el universo completo." % TOP_FILAS,
            "Cobertura por año: 2023-2025 = adjudicaciones SEACE; 2026 = órdenes Perú Compras (el consolidado SEACE del año en curso aún no se publica).",
            "El año en curso corresponde al acumulado disponible hasta la última actualización.",
            "Las alertas son señales estadísticas y no constituyen prueba de irregularidad."],
        "log_actualizacion": (manifest.get("log_actualizacion", []) + [
            {"fecha": hoy(), "evento": "Ingesta SEACE %s + MEF" % ", ".join(map(str, sorted(aggs))), "estado": "ok"}])[-30:]})
    guardar_json("manifest.json", manifest)
    log("LISTO. Entidades: %d | Proveedores(top): %d | Años SEACE: %s" % (len(ents), len(top_prov_global), sorted(aggs)))
    for e in evol:
        log("  %s: adjudicado=%s pim=%s" % (e["anio"], e["adjudicado"], e["pim"]))
    return 0

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--anios", nargs="+", type=int, default=[2023, 2024, 2025])
    p.add_argument("--parcial", type=int, default=2026)
    a = p.parse_args()
    sys.exit(main(a.anios, a.parcial))
